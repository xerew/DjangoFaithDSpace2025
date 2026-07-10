import base64
import io
import os
import re
import zipfile
from pathlib import Path

from django.conf import settings
from django.utils.html import strip_tags
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    Simulation, ExperimentLL, VRARExperiment, Subject,
    NextQuestionLogic, QuestionBunch, EvQuestionBranching,
)
from .template_generator import generate_blank_template


_MIME_TO_EXT = {
    'image/png': '.png',
    'image/jpeg': '.jpg',
    'image/jpg': '.jpg',
    'image/gif': '.gif',
    'image/webp': '.webp',
    'image/svg+xml': '.svg',
    'image/bmp': '.bmp',
}


def _rewrite_images(html, prefix):
    """
    Find every <img src="..."> in html and replace it with a ZIP-relative path.

    Handles:
      - base64 data URIs  (data:image/png;base64,...)
      - server media URLs (/media/... or MEDIA_URL/...)

    Returns:
      (rewritten_html, image_list)

    image_list entries are one of:
      ('file',  zip_name, abs_path)   – copy an existing file into the ZIP
      ('bytes', zip_name, img_bytes)  – write raw bytes into the ZIP
    """
    image_list = []
    counter = [0]
    media_url = getattr(settings, 'MEDIA_URL', '/media/').rstrip('/')

    def _replace(m):
        src = m.group(1)

        # ── base64 data URI ────────────────────────────────────────────────
        dm = re.match(r'data:(image/[^;]+);base64,(.+)', src, re.DOTALL)
        if dm:
            mime = dm.group(1).lower()
            ext = _MIME_TO_EXT.get(mime, '.jpg')
            try:
                img_bytes = base64.b64decode(dm.group(2))
            except Exception:
                return m.group(0)
            counter[0] += 1
            zip_name = f'{prefix}_image_{counter[0]}{ext}'
            image_list.append(('bytes', zip_name, img_bytes))
            return f'src="images/{zip_name}"'

        # ── server media URL ───────────────────────────────────────────────
        if src.startswith(media_url + '/'):
            rel = src[len(media_url) + 1:]
        elif src.startswith('/media/'):
            rel = src[len('/media/'):]
        else:
            return m.group(0)

        abs_path = os.path.join(settings.MEDIA_ROOT, rel)
        if not os.path.isfile(abs_path):
            return m.group(0)
        counter[0] += 1
        ext = Path(abs_path).suffix.lower() or '.jpg'
        zip_name = f'{prefix}_image_{counter[0]}{ext}'
        image_list.append(('file', zip_name, abs_path))
        return f'src="images/{zip_name}"'

    rewritten = re.sub(r'src="([^"]*)"', _replace, html, flags=re.DOTALL)
    return rewritten, image_list


def _col_map(ws):
    return {
        str(cell.value).replace('*', '').strip(): cell.column
        for cell in ws[1] if cell.value
    }


def _write_row(ws, cols, row_num, data):
    for col_name, value in data.items():
        col_idx = cols.get(col_name)
        if col_idx is not None:
            ws.cell(row=row_num, column=col_idx, value=value if value is not None else '')


class ScenarioExporter:
    def __init__(self, scenario):
        self.scenario = scenario

    def build_workbook_and_images(self):
        """Returns (wb, image_list) where each entry is ('file', zip_name, abs_path)
        or ('bytes', zip_name, img_bytes)."""
        s = self.scenario

        simulations = list(Simulation.objects.values_list('name', flat=True).order_by('name'))
        remote_labs = list(ExperimentLL.objects.values_list('name', flat=True).order_by('name'))
        vr_labs = list(VRARExperiment.objects.values_list('name', flat=True).order_by('name'))
        subjects_qs = list(Subject.objects.values_list('name', flat=True).order_by('name'))

        wb = generate_blank_template(simulations, remote_labs, vr_labs, subjects_qs)

        all_images = []

        # ── Scenario cover image ───────────────────────────────────────────
        if s.image and s.image.name:
            abs_path = os.path.join(settings.MEDIA_ROOT, s.image.name)
            if os.path.isfile(abs_path):
                ext = Path(abs_path).suffix.lower() or '.jpg'
                all_images.append(('file', f'scenario_cover{ext}', abs_path))

        # ── Scenario sheet ─────────────────────────────────────────────────
        ws_s = wb['Scenario']
        cols_s = _col_map(ws_s)
        subj_list = list(s.subjects.values_list('name', flat=True)[:3])
        age_min = s.age_of_students.lower if s.age_of_students else None
        age_max = s.age_of_students.upper if s.age_of_students else None
        _write_row(ws_s, cols_s, 2, {
            'Name': s.name,
            'Description': s.description or '',
            'Learning Goals': s.learning_goals or '',
            'Language': s.language or '',
            'Age Min': age_min,
            'Age Max': age_max,
            'Suggested Time (min)': s.suggested_learning_time,
            'Visibility': s.visibility_status or 'private',
            'Subject 1': subj_list[0] if len(subj_list) > 0 else '',
            'Subject 2': subj_list[1] if len(subj_list) > 1 else '',
            'Subject 3': subj_list[2] if len(subj_list) > 2 else '',
        })

        # ── Phases ─────────────────────────────────────────────────────────
        ws_ph = wb['Phases']
        cols_ph = _col_map(ws_ph)
        phases = list(s.phases.order_by('created_on'))
        for row_num, ph in enumerate(phases, start=2):
            _write_row(ws_ph, cols_ph, row_num, {
                'Phase Name': ph.name,
                'Description': ph.description or '',
            })

        # ── Activities ─────────────────────────────────────────────────────
        ws_acts = wb['Activities']
        cols_acts = _col_map(ws_acts)

        # Append hidden "Text HTML" column
        html_col_idx = max(cols_acts.values()) + 1
        hdr = ws_acts.cell(row=1, column=html_col_idx, value='Text HTML')
        hdr.font = Font(bold=True, color='FFFFFF')
        hdr.fill = PatternFill('solid', fgColor='6B7280')
        hdr.alignment = Alignment(horizontal='center')
        ws_acts.column_dimensions[hdr.column_letter].width = 20
        cols_acts['Text HTML'] = html_col_idx

        activities = list(
            s.activities.order_by('phase__created_on', 'created_on')
            .select_related('phase', 'activity_type', 'simulation', 'experiment_ll', 'vr_ar_experiment')
        )

        # Build answer key map: answer.pk -> 'ans_N' (per activity)
        answer_key_map = {}
        for act in activities:
            for i, ans in enumerate(act.answers.order_by('created_on'), start=1):
                answer_key_map[ans.pk] = f'ans_{i}'

        for row_num, act in enumerate(activities, start=2):
            phase_slug = re.sub(r'[^\w]+', '_', (act.phase.name or '').strip()).strip('_').lower()
            act_slug = re.sub(r'[^\w]+', '_', (act.name or '').strip()).strip('_').lower()
            img_prefix = f'phase_{phase_slug}_activity_{act_slug}'

            rewritten_html, img_list = _rewrite_images(act.text or '', img_prefix)
            all_images.extend(img_list)

            exp_type = sim_name = rlab_name = vr_name = ''
            if act.simulation_id:
                exp_type = 'Simulation'
                sim_name = act.simulation.name
            elif act.experiment_ll_id:
                exp_type = 'Remote Lab'
                rlab_name = act.experiment_ll.name
            elif act.vr_ar_experiment_id:
                exp_type = 'VR/AR Lab'
                vr_name = act.vr_ar_experiment.name

            _write_row(ws_acts, cols_acts, row_num, {
                'Activity Name': act.name,
                'Phase Name': act.phase.name if act.phase else '',
                'Text': act.plain_text or strip_tags(act.text),
                'Activity Type': act.activity_type.name if act.activity_type else '',
                'Helper': act.helper or '',
                'Is Evaluatable': 'Yes' if act.is_evaluatable else 'No',
                'Is Primary Evaluation': 'Yes' if act.is_primary_ev else 'No',
                'Experiment Type': exp_type,
                'Simulation Name': sim_name,
                'Remote Lab Name': rlab_name,
                'VR Lab Name': vr_name,
                'Text HTML': rewritten_html or '',
            })

        # ── Answers ────────────────────────────────────────────────────────
        ws_ans = wb['Answers']
        cols_ans = _col_map(ws_ans)
        ans_row = 2
        for act in activities:
            for i, ans in enumerate(act.answers.order_by('created_on'), start=1):
                _write_row(ws_ans, cols_ans, ans_row, {
                    'Activity Name': act.name,
                    'Answer Key': f'ans_{i}',
                    'Answer Text': ans.text,
                    'Is Correct': 'Yes' if ans.is_correct else 'No',
                    'Answer Weight': '' if ans.is_correct else str(ans.answer_weight),
                })
                ans_row += 1

        # ── Next Activity ──────────────────────────────────────────────────
        ws_next = wb['Next Activity']
        cols_next = _col_map(ws_next)
        # Source Phase / Next Phase columns are already in the blank template
        # with VLOOKUP formulas; exporter overwrites data rows with actual strings.

        routings = (
            NextQuestionLogic.objects
            .filter(activity__scenario=s)
            .select_related('activity__phase', 'answer', 'next_activity__phase')
            .order_by('activity__phase__created_on', 'activity__created_on', 'id')
        )
        # Deduplicate on (phase_name, activity_name, ans_key) — keeps first occurrence
        # which matches how the platform resolves the active routing record.
        seen_routing = set()
        row_num = 2
        for r in routings:
            ans_key = answer_key_map.get(r.answer_id, '') if r.answer_id else ''
            src_phase = r.activity.phase.name if r.activity.phase else ''
            pair = (src_phase, r.activity.name, ans_key)
            if pair in seen_routing:
                continue
            seen_routing.add(pair)
            # Source Phase / Next Phase are left as VLOOKUP formulas from the
            # template so they auto-update when teachers edit the file in Excel.
            _write_row(ws_next, cols_next, row_num, {
                'Source Activity Name': r.activity.name,
                'Answer Key': ans_key,
                'Next Activity Name': r.next_activity.name if r.next_activity else '',
            })
            row_num += 1

        # ── Evaluation ─────────────────────────────────────────────────────
        ws_eval = wb['Evaluation']
        cols_eval = _col_map(ws_eval)
        act_id_to_name = {act.id: act.name for act in activities}

        bunches = (
            QuestionBunch.objects
            .filter(activity_primary__scenario=s)
            .select_related('activity_primary')
            .order_by('activity_primary__phase__created_on', 'activity_primary__created_on')
        )
        branchings = {
            b.activity_id: b
            for b in EvQuestionBranching.objects
            .filter(activity__scenario=s)
            .select_related('next_question_on_high', 'next_question_on_mid', 'next_question_on_low')
        }
        for row_num, bunch in enumerate(bunches, start=2):
            primary = bunch.activity_primary
            grouped_names = [act_id_to_name.get(aid, '') for aid in bunch.activity_ids]
            branching = branchings.get(primary.pk)
            high = branching.next_question_on_high.name if branching and branching.next_question_on_high else ''
            mid = branching.next_question_on_mid.name if branching and branching.next_question_on_mid else ''
            low = branching.next_question_on_low.name if branching and branching.next_question_on_low else ''
            row_data = {'Primary Activity Name': primary.name}
            for i, gname in enumerate(grouped_names[:6], start=1):
                row_data[f'Grouped Activity {i}'] = gname
            row_data['High Performers Activity'] = high
            row_data['Moderate Performers Activity'] = mid
            row_data['Low Performers Activity'] = low
            _write_row(ws_eval, cols_eval, row_num, row_data)

        return wb, all_images

    def to_zip_bytes(self, xlsx_name):
        wb, all_images = self.build_workbook_and_images()

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_buf.seek(0)

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(xlsx_name, xlsx_buf.read())
            for item in all_images:
                kind, zip_name = item[0], item[1]
                if kind == 'file':
                    zf.write(item[2], f'images/{zip_name}')
                else:
                    zf.writestr(f'images/{zip_name}', item[2])

        zip_buf.seek(0)
        return zip_buf.read()
