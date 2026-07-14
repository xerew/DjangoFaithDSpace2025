import csv
import io
import json
import openpyxl

from django.contrib.auth.models import User, Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import TestCase, Client
from django.urls import reverse

from authoringtool.models import (
    Activity,
    ActivityProposal,
    ActivityProposalEditEvent,
    ActivityType,
    Answer,
    AnswerFeedback,
    NextQuestionLogic,
    Phase,
    ProposalGenerationRun,
    Scenario,
    SchoolDepartment,
    UserAnswer,
    UserProposalReview,
    UserScenarioScore,
)
from authoringtool.tasks import compute_student_performance_metrics


def make_xlsx(
    scenario_name='Test Scenario',
    phases=None,
    activities=None,
    answers=None,
    routing=None,
    evaluation=None,
    missing_sheet=None,
    activity_type='Explanation',
):
    """Build a minimal valid .xlsx in memory. Each argument is a list of tuples (row values)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'README'
    ws['A1'] = 'Instructions'

    sheets = {
        'Scenario': (
            ['Name', 'Description', 'Visibility'],
            [[scenario_name, '', 'private']],
        ),
        'Phases': (
            ['Phase Name', 'Description'],
            phases or [['Phase 1', '']],
        ),
        'Activities': (
            ['Activity Name', 'Phase Name', 'Text', 'Activity Type', 'Helper',
             'Is Evaluatable', 'Is Primary Evaluation',
             'Experiment Type', 'Simulation Name', 'Remote Lab Name', 'VR Lab Name'],
            activities or [['Act 1', 'Phase 1', 'Hello world', activity_type, '', 'No', 'No', '', '', '', '']],
        ),
        'Answers': (
            ['Activity Name', 'Answer Key', 'Answer Text', 'Is Correct', 'Answer Weight'],
            answers or [],
        ),
        'Next Activity': (
            ['Source Activity Name', 'Answer Key', 'Next Activity Name'],
            routing or [],
        ),
        'Evaluation': (
            ['Primary Activity Name',
             'Grouped Activity 1', 'Grouped Activity 2', 'Grouped Activity 3',
             'Grouped Activity 4', 'Grouped Activity 5', 'Grouped Activity 6',
             'High Performers Activity', 'Moderate Performers Activity', 'Low Performers Activity'],
            evaluation or [],
        ),
    }

    for name, (header, rows) in sheets.items():
        if name == missing_sheet:
            continue
        ws2 = wb.create_sheet(name)
        ws2.append(header)
        for row in rows:
            ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ImporterValidationTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('teacher', password='pass')
        ActivityType.objects.create(name='Explanation', created_by=self.user, updated_by=self.user)
        ActivityType.objects.create(name='Question', created_by=self.user, updated_by=self.user)

    def _run(self, **kwargs):
        from authoringtool.importer import ScenarioImporter
        buf = make_xlsx(**kwargs)
        importer = ScenarioImporter(buf, self.user)
        importer._parse()
        if not importer.errors:
            importer._validate()
        return importer.errors

    def test_valid_file_no_errors(self):
        errors = self._run()
        self.assertEqual(errors, [])

    def test_missing_sheet_reported(self):
        errors = self._run(missing_sheet='Activities')
        messages = [e['message'] for e in errors]
        self.assertTrue(any('Activities' in m for m in messages))

    def test_missing_scenario_name(self):
        errors = self._run(scenario_name='')
        messages = [e['message'] for e in errors]
        self.assertTrue(any('Name' in m or 'required' in m.lower() for m in messages))

    def test_duplicate_scenario_name(self):
        # Scenario.objects.create() fails on SQLite because Scenario.age_of_students is an
        # IntegerRangeField (postgres-only).  Mock the DB lookup instead.
        from unittest.mock import patch
        from authoringtool.importer import ScenarioImporter
        buf = make_xlsx(scenario_name='Test Scenario')
        importer = ScenarioImporter(buf, self.user)
        importer._parse()
        with patch('authoringtool.importer.Scenario.objects') as mock_qs:
            mock_qs.filter.return_value.exists.return_value = True
            importer._load_db_lookups()
            importer._validate_scenario()
        messages = [e['message'] for e in importer.errors]
        self.assertTrue(any('already exists' in m for m in messages))

    def test_invalid_visibility(self):
        # build with bad visibility by patching scenario_data after parse
        from authoringtool.importer import ScenarioImporter
        buf = make_xlsx()
        importer = ScenarioImporter(buf, self.user)
        importer._parse()
        importer.scenario_data['Visibility'] = 'invalid'
        importer._load_db_lookups()
        importer._validate_scenario()
        self.assertTrue(any('Visibility' in e['column'] for e in importer.errors))

    def test_activity_references_unknown_phase(self):
        errors = self._run(
            activities=[['Act 1', 'NonexistentPhase', 'Hello', 'Explanation', '', 'No', 'No', '', '', '']],
        )
        self.assertTrue(any('Phase' in e['message'] for e in errors))

    def test_duplicate_activity_names(self):
        errors = self._run(
            activities=[
                ['Act 1', 'Phase 1', 'Hello', 'Explanation', '', 'No', 'No', '', '', ''],
                ['Act 1', 'Phase 1', 'Dupe',  'Explanation', '', 'No', 'No', '', '', ''],
            ],
        )
        self.assertTrue(any('Duplicate' in e['message'] for e in errors))

    def test_primary_ev_without_evaluatable(self):
        errors = self._run(
            activities=[['Act 1', 'Phase 1', 'Hello', 'Explanation', '', 'No', 'Yes', '', '', '']],
        )
        self.assertTrue(any('Is Evaluatable' in e['message'] or 'Is Primary' in e['message'] for e in errors))

    def test_evaluatable_without_evaluation_row(self):
        errors = self._run(
            activities=[['Quiz', 'Phase 1', 'Q?', 'Question', '', 'Yes', 'Yes', '', '', '']],
            answers=[['Quiz', 'ans_a', 'Option A', 'Yes', '1']],
            evaluation=[],
        )
        self.assertTrue(any('Evaluation' in e['message'] or 'evaluatable' in e['message'].lower() for e in errors))

    def test_answer_references_unknown_activity(self):
        errors = self._run(
            answers=[['NonexistentAct', 'ans_a', 'Option A', 'No', '']],
        )
        self.assertTrue(any('NonexistentAct' in e['message'] for e in errors))

    def test_routing_references_unknown_next_activity(self):
        errors = self._run(
            routing=[['Act 1', '', 'GhostActivity']],
        )
        self.assertTrue(any('GhostActivity' in e['message'] for e in errors))

    def test_routing_answer_not_found(self):
        errors = self._run(
            answers=[['Act 1', 'ans_a', 'Option A', 'Yes', '1']],
            routing=[['Act 1', 'wrong_key', '']],
        )
        self.assertTrue(any('wrong_key' in e['message'] for e in errors))

    def test_invalid_boolean_field(self):
        errors = self._run(
            activities=[['Act 1', 'Phase 1', 'Hello', 'Explanation', '', 'maybe', 'No', '', '', '']],
        )
        self.assertTrue(any('Yes' in e['message'] or 'No' in e['message'] for e in errors))


class ImporterCreationTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('teacher2', password='pass')
        ActivityType.objects.create(name='Explanation', created_by=self.user, updated_by=self.user)
        ActivityType.objects.create(name='Question', created_by=self.user, updated_by=self.user)

    def _import(self, **kwargs):
        from authoringtool.importer import ScenarioImporter
        buf = make_xlsx(**kwargs)
        importer = ScenarioImporter(buf, self.user)
        return importer.run()

    def test_creates_scenario(self):
        scenario, errors = self._import(scenario_name='Created Scenario')
        self.assertEqual(errors, [])
        self.assertIsNotNone(scenario)
        self.assertTrue(Scenario.objects.filter(name='Created Scenario').exists())

    def test_creates_phases(self):
        scenario, errors = self._import(
            phases=[['Phase A', 'First phase'], ['Phase B', '']],
            activities=[
                ['Act 1', 'Phase A', 'Hello', 'Explanation', '', 'No', 'No', '', '', ''],
                ['Act 2', 'Phase B', 'World', 'Explanation', '', 'No', 'No', '', '', ''],
            ],
        )
        self.assertEqual(errors, [])
        self.assertEqual(Phase.objects.filter(scenario=scenario).count(), 2)
        phase_names = list(Phase.objects.filter(scenario=scenario).values_list('name', flat=True))
        self.assertIn('Phase A', phase_names)
        self.assertIn('Phase B', phase_names)

    def test_creates_activities_with_correct_phase(self):
        scenario, errors = self._import(
            activities=[['MyAct', 'Phase 1', '**Bold text**', 'Explanation', '', 'No', 'No', '', '', '']],
        )
        self.assertEqual(errors, [])
        act = Activity.objects.get(scenario=scenario)
        self.assertEqual(act.name, 'MyAct')
        self.assertIn('<strong>', act.text)  # markdown converted
        self.assertNotIn('<strong>', act.plain_text)  # tags stripped

    def test_creates_answers(self):
        scenario, errors = self._import(
            answers=[
                ['Act 1', 'ans_correct', 'Option A', 'Yes', '2'],
                ['Act 1', 'ans_wrong',   'Option B', 'No',  '2'],
            ],
        )
        self.assertEqual(errors, [])
        act = Activity.objects.get(scenario=scenario)
        self.assertEqual(act.answers.count(), 2)
        correct = act.answers.get(is_correct=True)
        self.assertEqual(correct.answer_weight, 3)  # auto-set to 3 regardless of sheet value
        wrong = act.answers.get(is_correct=False)
        self.assertEqual(wrong.answer_weight, 2)    # sheet value honoured for wrong answers

    def test_creates_next_question_logic_default(self):
        scenario, errors = self._import(
            activities=[
                ['Act 1', 'Phase 1', 'First',  'Explanation', '', 'No', 'No', '', '', ''],
                ['Act 2', 'Phase 1', 'Second', 'Explanation', '', 'No', 'No', '', '', ''],
            ],
            routing=[['Act 1', '', 'Act 2']],
        )
        self.assertEqual(errors, [])
        act1 = Activity.objects.get(scenario=scenario, name='Act 1')
        act2 = Activity.objects.get(scenario=scenario, name='Act 2')
        logic = NextQuestionLogic.objects.get(activity=act1, answer__isnull=True)
        self.assertEqual(logic.next_activity, act2)

    def test_creates_next_question_logic_per_answer(self):
        scenario, errors = self._import(
            activities=[
                ['Q',    'Phase 1', 'Question?',  'Question',    '', 'No', 'No', '', '', ''],
                ['Good', 'Phase 1', 'Well done!', 'Explanation', '', 'No', 'No', '', '', ''],
                ['Bad',  'Phase 1', 'Try again!', 'Explanation', '', 'No', 'No', '', '', ''],
            ],
            answers=[
                ['Q', 'ans_correct', 'Correct', 'Yes', '1'],
                ['Q', 'ans_wrong',   'Wrong',   'No',  '0'],
            ],
            routing=[
                ['Q', 'ans_correct', 'Good'],
                ['Q', 'ans_wrong',   'Bad'],
            ],
        )
        self.assertEqual(errors, [])
        q_act = Activity.objects.get(scenario=scenario, name='Q')
        correct_ans = q_act.answers.get(is_correct=True)
        logic = NextQuestionLogic.objects.get(activity=q_act, answer=correct_ans)
        self.assertEqual(logic.next_activity.name, 'Good')

    def test_created_by_is_set(self):
        scenario, errors = self._import()
        self.assertEqual(errors, [])
        self.assertEqual(scenario.created_by, self.user)
        self.assertEqual(scenario.updated_by, self.user)


class PerActivityCSVColumnTest(TestCase):
    """
    Verify that compute_student_performance_metrics with include_activity_detail=True
    produces a CSV whose header contains the expected per-activity column names and
    whose header length matches every data row.
    """

    def setUp(self):
        # --- Admin / default user required by SET_DEFAULT FKs ---
        self.admin = User.objects.create_user(username="admin", password="admin")

        # --- School department so the user passes the empty-group_ids filter ---
        dept = SchoolDepartment.objects.create(name="Test Dept")

        # --- Student user ---
        self.student = User.objects.create_user(
            username="student1", password="pass"
        )
        self.student.school_department = dept
        self.student.save()

        # --- Scenario ---
        self.scenario = Scenario.objects.create(
            name="Test Scenario",
            suggested_learning_time=60,
            created_by=self.admin,
            updated_by=self.admin,
        )

        # --- Phase ---
        self.phase = Phase.objects.create(
            name="Phase 1",
            scenario=self.scenario,
            created_by=self.admin,
            updated_by=self.admin,
        )

        # --- Activity types ---
        at_read = ActivityType.objects.create(
            name="Read Text",
            created_by=self.admin,
            updated_by=self.admin,
        )
        at_quiz = ActivityType.objects.create(
            name="Quiz",
            created_by=self.admin,
            updated_by=self.admin,
        )

        # --- Activities ---
        # Non-evaluatable reading activity
        self.act_read = Activity.objects.create(
            name="Read Text",
            text="Read this passage.",
            is_evaluatable=False,
            is_primary_ev=False,
            phase=self.phase,
            scenario=self.scenario,
            activity_type=at_read,
            created_by=self.admin,
            updated_by=self.admin,
        )
        # Evaluatable quiz activity (primary evaluatable)
        self.act_quiz = Activity.objects.create(
            name="Quiz",
            text="Answer this question.",
            is_evaluatable=True,
            is_primary_ev=True,
            phase=self.phase,
            scenario=self.scenario,
            activity_type=at_quiz,
            created_by=self.admin,
            updated_by=self.admin,
        )

        # --- Answers for the quiz activity ---
        self.answer_correct = Answer.objects.create(
            activity=self.act_quiz,
            text="Correct answer",
            is_correct=True,
            answer_weight=10,
            created_by=self.admin,
            updated_by=self.admin,
        )
        Answer.objects.create(
            activity=self.act_quiz,
            text="Wrong answer",
            is_correct=False,
            answer_weight=0,
            created_by=self.admin,
            updated_by=self.admin,
        )

        # --- UserScenarioScore: required for empty group_ids lookup ---
        UserScenarioScore.objects.create(
            user=self.student,
            scenario=self.scenario,
            user_score=10,
        )

        # --- UserAnswers ---
        UserAnswer.objects.create(
            user=self.student,
            activity=self.act_read,
            answer=None,
            timing=30,
        )
        UserAnswer.objects.create(
            user=self.student,
            activity=self.act_quiz,
            answer=self.answer_correct,
            timing=45,
        )

    def test_per_activity_header_names_and_column_count(self):
        result = compute_student_performance_metrics(
            scenario_id=self.scenario.id,
            group_ids=[],
            start_date=None,
            end_date=None,
            include_activity_detail=True,
        )

        self.assertNotIn("error", result, msg=f"Task returned error: {result}")
        self.assertIn("csv_content", result)

        reader = csv.reader(io.StringIO(result["csv_content"]))
        rows = list(reader)
        self.assertGreaterEqual(len(rows), 2, "Expected at least a header and one data row")

        header = rows[0]
        data_row = rows[1]

        # --- Assert per-activity column names ---
        # Activities are sorted by id; act_read was created first so it comes first.
        expected_cols = [
            "Phase 1 > Read Text Type",
            "Phase 1 > Read Text Time (s)",
            "Phase 1 > Read Text Score",
            "Phase 1 > Quiz Type",
            "Phase 1 > Quiz Time (s)",
            "Phase 1 > Quiz Score",
        ]
        for col in expected_cols:
            self.assertIn(
                col,
                header,
                msg=f"Expected column '{col}' not found in header: {header}",
            )

        # --- Assert header and data row lengths match ---
        self.assertEqual(
            len(header),
            len(data_row),
            msg=(
                f"Header has {len(header)} columns but data row has {len(data_row)} columns.\n"
                f"Header: {header}\nData: {data_row}"
            ),
        )

        # --- Assert Quiz Score cell value ---
        quiz_score_idx = header.index("Phase 1 > Quiz Score")
        self.assertEqual(data_row[quiz_score_idx], "10", "Quiz score should equal the correct answer weight")


class TemplateDownloadTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user('teacher3', password='pass')
        g, _ = Group.objects.get_or_create(name='teachers')
        self.user.groups.add(g)
        self.client.login(username='teacher3', password='pass')

    def test_download_requires_login(self):
        self.client.logout()
        r = self.client.get(reverse('download_template'))
        self.assertNotEqual(r.status_code, 200)

    def test_download_returns_xlsx(self):
        r = self.client.get(reverse('download_template'))
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])

    def test_download_has_all_sheets(self):
        import openpyxl, io
        r = self.client.get(reverse('download_template'))
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        for sheet in ['README', 'Scenario', 'Phases', 'Activities', 'Answers', 'Next Activity', 'Evaluation']:
            self.assertIn(sheet, wb.sheetnames)


class ImportViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user('teacher4', password='pass')
        g, _ = Group.objects.get_or_create(name='teachers')
        self.user.groups.add(g)
        self.client.login(username='teacher4', password='pass')
        ActivityType.objects.create(name='Explanation', created_by=self.user, updated_by=self.user)
        ActivityType.objects.create(name='Question', created_by=self.user, updated_by=self.user)

    def _upload(self, buf, filename='template.xlsx'):
        f = SimpleUploadedFile(filename, buf.read(),
                               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return self.client.post(reverse('import_scenario'), {'template_file': f})

    def test_requires_login(self):
        self.client.logout()
        buf = make_xlsx(scenario_name='X')
        r = self._upload(buf)
        self.assertNotEqual(r.status_code, 200)

    def test_valid_file_creates_scenario(self):
        buf = make_xlsx(scenario_name='Imported Scenario')
        r = self._upload(buf)
        data = json.loads(r.content)
        self.assertTrue(data['success'])
        self.assertTrue(Scenario.objects.filter(name='Imported Scenario').exists())

    def test_valid_file_returns_redirect_url(self):
        buf = make_xlsx(scenario_name='Redir Test')
        r = self._upload(buf)
        data = json.loads(r.content)
        self.assertIn('redirect', data)
        self.assertIn(str(data['scenario_id']), data['redirect'])

    def test_invalid_file_returns_errors(self):
        buf = make_xlsx(scenario_name='')  # missing required name
        r = self._upload(buf)
        data = json.loads(r.content)
        self.assertFalse(data['success'])
        self.assertIsInstance(data['errors'], list)
        self.assertGreater(len(data['errors']), 0)

    def test_non_xlsx_rejected(self):
        f = SimpleUploadedFile('file.csv', b'name,val', content_type='text/csv')
        r = self.client.post(reverse('import_scenario'), {'template_file': f})
        data = json.loads(r.content)
        self.assertFalse(data['success'])

    def test_get_not_allowed(self):
        r = self.client.get(reverse('import_scenario'))
        self.assertEqual(r.status_code, 405)

    def test_duplicate_scenario_name_returns_error(self):
        Scenario.objects.create(name='Dupe', created_by=self.user, updated_by=self.user)
        buf = make_xlsx(scenario_name='Dupe')
        r = self._upload(buf)
        data = json.loads(r.content)
        self.assertFalse(data['success'])
        self.assertTrue(any('already exists' in e['message'] for e in data['errors']))


class ActivityProposalEditEventModelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('teacher_edit1', password='pass')
        self.scenario = Scenario.objects.create(
            name='Edit Event Scenario', created_by=self.user, updated_by=self.user
        )
        self.phase = Phase.objects.create(
            name='Phase 1', scenario=self.scenario, created_by=self.user, updated_by=self.user
        )
        self.activity_type = ActivityType.objects.create(
            name='Explanation', created_by=self.user, updated_by=self.user
        )
        self.activity = Activity.objects.create(
            name='Act 1', text='Hello', scenario=self.scenario, phase=self.phase,
            activity_type=self.activity_type, created_by=self.user, updated_by=self.user,
        )
        self.proposal = ActivityProposal.objects.create(
            scenario=self.scenario, phase=self.phase, activity=self.activity,
            proposal_type='revise', suggested_action='raw', translated_action='raw',
            json_action=json.dumps({
                "activity_name": "Act 1", "content": "Old content",
                "explanation": "Old exp", "answers": [],
            }),
            json_translated_action=json.dumps({
                "activity_name": "Act 1", "content": "Old content",
                "explanation": "Old exp", "answers": [],
            }),
        )
        self.review = UserProposalReview.objects.create(proposal=self.proposal, user=self.user)

    def test_review_defaults(self):
        self.assertFalse(self.review.was_edited)
        self.assertEqual(self.review.edit_count, 0)

    def test_create_edit_event(self):
        event = ActivityProposalEditEvent.objects.create(
            review=self.review, edit_number=1,
            edited_json={
                "activity_name": "Act 1", "content": "New content",
                "explanation": "Old exp", "answers": [],
            },
            changed_fields={"content": {"changed": True, "char_delta": 3}},
        )
        self.assertEqual(self.review.edit_events.count(), 1)
        self.assertEqual(event.edit_number, 1)

    def test_unique_edit_number_per_review(self):
        ActivityProposalEditEvent.objects.create(
            review=self.review, edit_number=1, edited_json={}, changed_fields={},
        )
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                ActivityProposalEditEvent.objects.create(
                    review=self.review, edit_number=1, edited_json={}, changed_fields={},
                )


class EditProposalJsonViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user('teacher_edit2', password='pass')
        g, _ = Group.objects.get_or_create(name='teachers')
        self.user.groups.add(g)
        self.client.login(username='teacher_edit2', password='pass')

        self.scenario = Scenario.objects.create(
            name='Edit View Scenario', created_by=self.user, updated_by=self.user
        )
        self.phase = Phase.objects.create(
            name='Phase 1', scenario=self.scenario, created_by=self.user, updated_by=self.user
        )
        self.activity_type = ActivityType.objects.create(
            name='Explanation', created_by=self.user, updated_by=self.user
        )
        self.activity = Activity.objects.create(
            name='Act 1', text='Hello', scenario=self.scenario, phase=self.phase,
            activity_type=self.activity_type, created_by=self.user, updated_by=self.user,
        )
        self.proposal = ActivityProposal.objects.create(
            scenario=self.scenario, phase=self.phase, activity=self.activity,
            proposal_type='revise', suggested_action='raw', translated_action='raw',
            json_action=json.dumps({
                "activity_name": "Act 1",
                "content": "Original content",
                "explanation": "Original explanation",
                "answers": [{"text": "A. Old answer", "is_correct": True, "weight": 3}],
            }),
            json_translated_action='',
        )

    def _post_edit(self, **overrides):
        data = {
            'activity_name': 'Act 1',
            'content': 'Original content',
            'explanation': 'Original explanation',
            'answer_text_1': 'A. Old answer',
        }
        data.update(overrides)
        url = reverse('edit_proposal_json', args=[self.scenario.id, self.proposal.id])
        return self.client.post(url, data)

    def test_first_edit_creates_event_diffed_against_original_proposal(self):
        self._post_edit(content='Revised content is longer now')

        review = UserProposalReview.objects.get(proposal=self.proposal, user=self.user)
        self.assertTrue(review.was_edited)
        self.assertEqual(review.edit_count, 1)

        events = list(review.edit_events.order_by('edit_number'))
        self.assertEqual(len(events), 1)
        event = events[0]
        self.assertEqual(event.edit_number, 1)
        self.assertTrue(event.changed_fields['content']['changed'])
        self.assertFalse(event.changed_fields['explanation']['changed'])
        self.assertFalse(event.changed_fields['answers']['changed'])

    def test_second_edit_diffs_against_previous_edit_not_original(self):
        self._post_edit(content='First revision')
        self._post_edit(content='First revision, refined further')

        review = UserProposalReview.objects.get(proposal=self.proposal, user=self.user)
        self.assertEqual(review.edit_count, 2)

        events = list(review.edit_events.order_by('edit_number'))
        self.assertEqual(len(events), 2)

        second_event = events[1]
        expected_delta = len('First revision, refined further') - len('First revision')
        self.assertEqual(second_event.changed_fields['content']['char_delta'], expected_delta)

    def test_answers_count_delta_tracks_added_answer(self):
        self._post_edit(answer_text_1='A. Old answer', answer_text_2='B. New second answer')

        review = UserProposalReview.objects.get(proposal=self.proposal, user=self.user)
        event = review.edit_events.get(edit_number=1)
        self.assertEqual(event.changed_fields['answers']['count_delta'], 1)
        self.assertTrue(event.changed_fields['answers']['changed'])

    def test_teacher_edited_json_still_holds_latest_state_only(self):
        self._post_edit(content='First revision')
        self._post_edit(content='Second revision')

        review = UserProposalReview.objects.get(proposal=self.proposal, user=self.user)
        self.assertEqual(review.teacher_edited_json['content'], 'Second revision')

    def test_get_request_does_not_create_edit_event(self):
        url = reverse('edit_proposal_json', args=[self.scenario.id, self.proposal.id])
        response = self.client.get(url)
        self.assertFalse(UserProposalReview.objects.filter(proposal=self.proposal, user=self.user).exists())


class ProposalGenerationRunModelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('run_owner', password='pass')
        self.scenario = Scenario.objects.create(
            name='Run Scenario', created_by=self.user, updated_by=self.user
        )

    def test_start_new_creates_current_run(self):
        run = ProposalGenerationRun.start_new(self.scenario, self.user)
        self.assertTrue(run.is_current)
        self.assertEqual(run.scenario, self.scenario)
        self.assertEqual(run.created_by, self.user)

    def test_start_new_archives_previous_current_run(self):
        first_run = ProposalGenerationRun.start_new(self.scenario, self.user)
        second_run = ProposalGenerationRun.start_new(self.scenario, self.user)

        first_run.refresh_from_db()
        self.assertFalse(first_run.is_current)
        self.assertTrue(second_run.is_current)

    def test_only_one_current_run_per_scenario(self):
        ProposalGenerationRun.start_new(self.scenario, self.user)
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                ProposalGenerationRun.objects.create(
                    scenario=self.scenario, created_by=self.user, is_current=True,
                )

    def test_different_scenarios_can_each_have_a_current_run(self):
        other_scenario = Scenario.objects.create(
            name='Other Run Scenario', created_by=self.user, updated_by=self.user
        )
        run1 = ProposalGenerationRun.start_new(self.scenario, self.user)
        run2 = ProposalGenerationRun.start_new(other_scenario, self.user)
        self.assertTrue(run1.is_current)
        self.assertTrue(run2.is_current)


class TriggerLlmContextTaskPermissionTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.owner = User.objects.create_user('gen_owner', password='pass')
        self.other_teacher = User.objects.create_user('gen_other', password='pass')
        self.admin = User.objects.create_user('gen_admin', password='pass', is_staff=True)
        g, _ = Group.objects.get_or_create(name='teachers')
        self.owner.groups.add(g)
        self.other_teacher.groups.add(g)
        self.admin.groups.add(g)
        self.scenario = Scenario.objects.create(
            name='Gen Perm Scenario', created_by=self.owner, updated_by=self.owner
        )

    def test_non_owner_teacher_forbidden(self):
        from unittest.mock import patch
        self.client.login(username='gen_other', password='pass')
        url = reverse('generate_llm_context', args=[self.scenario.id])
        with patch('authoringtool.views.generate_llm_context_for_scenario.delay') as mock_delay:
            response = self.client.post(url)
        self.assertEqual(response.status_code, 403)
        mock_delay.assert_not_called()

    def test_non_teacher_forbidden(self):
        from unittest.mock import patch
        non_teacher = User.objects.create_user('gen_notteacher', password='pass')
        self.client.login(username='gen_notteacher', password='pass')
        url = reverse('generate_llm_context', args=[self.scenario.id])
        with patch('authoringtool.views.generate_llm_context_for_scenario.delay') as mock_delay:
            response = self.client.post(url)
        self.assertEqual(response.status_code, 403)
        mock_delay.assert_not_called()

    def test_owner_can_trigger(self):
        from unittest.mock import patch
        self.client.login(username='gen_owner', password='pass')
        url = reverse('generate_llm_context', args=[self.scenario.id])
        with patch('authoringtool.views.generate_llm_context_for_scenario.delay') as mock_delay:
            mock_delay.return_value.id = 'fake-task-id'
            response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        mock_delay.assert_called_once_with(self.scenario.id, force_rebuild=False, triggered_by_id=self.owner.id)

    def test_admin_can_trigger(self):
        from unittest.mock import patch
        self.client.login(username='gen_admin', password='pass')
        url = reverse('generate_llm_context', args=[self.scenario.id])
        with patch('authoringtool.views.generate_llm_context_for_scenario.delay') as mock_delay:
            mock_delay.return_value.id = 'fake-task-id'
            response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        mock_delay.assert_called_once()


class ProposalListViewCurrentRunScopingTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user('scope_owner', password='pass')
        g, _ = Group.objects.get_or_create(name='teachers')
        self.user.groups.add(g)
        self.client.login(username='scope_owner', password='pass')

        self.scenario = Scenario.objects.create(
            name='Scope Scenario', created_by=self.user, updated_by=self.user
        )
        self.phase = Phase.objects.create(
            name='Phase 1', scenario=self.scenario, created_by=self.user, updated_by=self.user
        )
        self.activity_type = ActivityType.objects.create(
            name='Explanation', created_by=self.user, updated_by=self.user
        )
        self.activity = Activity.objects.create(
            name='Act 1', text='Hello', scenario=self.scenario, phase=self.phase,
            activity_type=self.activity_type, created_by=self.user, updated_by=self.user,
        )

        self.old_run = ProposalGenerationRun.start_new(self.scenario, self.user)
        self.old_proposal = ActivityProposal.objects.create(
            scenario=self.scenario, generation_run=self.old_run, phase=self.phase, activity=self.activity,
            proposal_type='revise', suggested_action='old', translated_action='old',
            json_action='{}', json_translated_action='{}',
        )

        self.current_run = ProposalGenerationRun.start_new(self.scenario, self.user)
        self.current_proposal = ActivityProposal.objects.create(
            scenario=self.scenario, generation_run=self.current_run, phase=self.phase, activity=self.activity,
            proposal_type='revise', suggested_action='current', translated_action='current',
            json_action='{}', json_translated_action='{}',
        )

    def test_proposal_list_only_shows_current_run(self):
        url = reverse('proposal_list', args=[self.scenario.id])
        response = self.client.get(url)
        proposals = list(response.context['proposals'])
        self.assertEqual(proposals, [self.current_proposal])

    def test_old_run_is_archived(self):
        self.old_run.refresh_from_db()
        self.current_run.refresh_from_db()
        self.assertFalse(self.old_run.is_current)
        self.assertTrue(self.current_run.is_current)


class AcceptedReviewsForPersonalScenarioScopingTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('accepted_scope_owner', password='pass')
        self.scenario = Scenario.objects.create(
            name='Accepted Scope Scenario', created_by=self.user, updated_by=self.user
        )
        self.phase = Phase.objects.create(
            name='Phase 1', scenario=self.scenario, created_by=self.user, updated_by=self.user
        )
        self.activity_type = ActivityType.objects.create(
            name='Explanation', created_by=self.user, updated_by=self.user
        )
        self.activity = Activity.objects.create(
            name='Act 1', text='Hello', scenario=self.scenario, phase=self.phase,
            activity_type=self.activity_type, created_by=self.user, updated_by=self.user,
        )

        self.old_run = ProposalGenerationRun.start_new(self.scenario, self.user)
        self.old_proposal = ActivityProposal.objects.create(
            scenario=self.scenario, generation_run=self.old_run, phase=self.phase, activity=self.activity,
            proposal_type='revise', suggested_action='old', translated_action='old',
            json_action='{}', json_translated_action='{}',
        )
        self.old_review = UserProposalReview.objects.create(
            proposal=self.old_proposal, user=self.user, status='accepted',
        )

        self.current_run = ProposalGenerationRun.start_new(self.scenario, self.user)
        self.current_proposal = ActivityProposal.objects.create(
            scenario=self.scenario, generation_run=self.current_run, phase=self.phase, activity=self.activity,
            proposal_type='revise', suggested_action='current', translated_action='current',
            json_action='{}', json_translated_action='{}',
        )
        self.current_review = UserProposalReview.objects.create(
            proposal=self.current_proposal, user=self.user, status='accepted',
        )

    def test_only_current_run_accepted_reviews_are_returned(self):
        from authoringtool.tasks import get_accepted_reviews_for_personal_scenario
        reviews = list(get_accepted_reviews_for_personal_scenario(self.scenario, self.user))
        self.assertEqual(reviews, [self.current_review])


class ProposalHistoryViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user('history_user', password='pass')
        g, _ = Group.objects.get_or_create(name='teachers')
        self.user.groups.add(g)
        self.client.login(username='history_user', password='pass')

        self.scenario = Scenario.objects.create(
            name='History Scenario', created_by=self.user, updated_by=self.user
        )
        self.phase = Phase.objects.create(
            name='Phase 1', scenario=self.scenario, created_by=self.user, updated_by=self.user
        )
        self.activity_type = ActivityType.objects.create(
            name='Explanation', created_by=self.user, updated_by=self.user
        )
        self.activity = Activity.objects.create(
            name='Act 1', text='Hello', scenario=self.scenario, phase=self.phase,
            activity_type=self.activity_type, created_by=self.user, updated_by=self.user,
        )

        self.old_run = ProposalGenerationRun.start_new(self.scenario, self.user)
        self.old_proposal_accepted = ActivityProposal.objects.create(
            scenario=self.scenario, generation_run=self.old_run, phase=self.phase, activity=self.activity,
            proposal_type='revise', suggested_action='Old accepted proposal', translated_action='x',
            json_action='{}', json_translated_action='{}',
        )
        UserProposalReview.objects.create(
            proposal=self.old_proposal_accepted, user=self.user, status='accepted',
        )
        self.old_proposal_rejected = ActivityProposal.objects.create(
            scenario=self.scenario, generation_run=self.old_run, phase=self.phase, activity=self.activity,
            proposal_type='skip', suggested_action='Old rejected proposal', translated_action='x',
            json_action='{}', json_translated_action='{}',
        )
        UserProposalReview.objects.create(
            proposal=self.old_proposal_rejected, user=self.user, status='rejected', rejection_reasons=['not relevant'],
        )

        self.current_run = ProposalGenerationRun.start_new(self.scenario, self.user)
        self.current_proposal = ActivityProposal.objects.create(
            scenario=self.scenario, generation_run=self.current_run, phase=self.phase, activity=self.activity,
            proposal_type='create', suggested_action='Current proposal', translated_action='x',
            json_action='{}', json_translated_action='{}',
        )

    def test_history_index_lists_only_past_runs(self):
        url = reverse('proposal_history', args=[self.scenario.id])
        response = self.client.get(url)
        run_summaries = response.context['run_summaries']
        self.assertEqual(len(run_summaries), 1)
        self.assertEqual(run_summaries[0]['run'], self.old_run)

    def test_history_index_shows_decision_counts(self):
        url = reverse('proposal_history', args=[self.scenario.id])
        response = self.client.get(url)
        summary = response.context['run_summaries'][0]
        self.assertEqual(summary['accepted'], 1)
        self.assertEqual(summary['rejected'], 1)
        self.assertEqual(summary['total'], 2)

    def test_run_detail_shows_that_runs_proposals_only(self):
        url = reverse('proposal_history_run_detail', args=[self.scenario.id, self.old_run.id])
        response = self.client.get(url)
        self.assertContains(response, 'Old accepted proposal')
        self.assertContains(response, 'Old rejected proposal')
        self.assertNotContains(response, 'Current proposal')

    def test_run_detail_shows_rejection_reasons(self):
        url = reverse('proposal_history_run_detail', args=[self.scenario.id, self.old_run.id])
        response = self.client.get(url)
        self.assertContains(response, 'not relevant')

    def test_run_from_wrong_scenario_404s(self):
        other_scenario = Scenario.objects.create(
            name='Other History Scenario', created_by=self.user, updated_by=self.user
        )
        url = reverse('proposal_history_run_detail', args=[other_scenario.id, self.old_run.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    def test_login_required_for_history(self):
        self.client.logout()
        url = reverse('proposal_history', args=[self.scenario.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)
