import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill('solid', fgColor='1D4ED8')
_REQ_FILL = PatternFill('solid', fgColor='DC2626')

_SCENARIO_LANGUAGES = (
    'English,Greek,Spanish,French,German,Italian,'
    'Portuguese,Dutch,Polish,Romanian,Turkish,Arabic,Other'
)

_GROUPED_ACT_COLS = [f'Grouped Activity {i}' for i in range(1, 7)]


def _write_sheet(wb, title, columns, bool_cols=(), example_rows=(), list_cols=None):
    """columns: list of (name, required_bool).
    list_cols: dict of {col_name: formula1} for dropdown validation.
    bool_cols: shorthand for Yes/No dropdowns (merged into list_cols internally)."""
    ws = wb.create_sheet(title)
    for col_idx, (name, required) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=f'{name}*' if required else name)
        cell.font = _HEADER_FONT
        cell.fill = _REQ_FILL if required else _HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(name) + 4, 16)

    col_name_to_idx = {name: i + 1 for i, (name, _) in enumerate(columns)}

    all_list_cols = {col: '"Yes,No"' for col in bool_cols}
    if list_cols:
        all_list_cols.update(list_cols)

    for col_name, formula1 in all_list_cols.items():
        if col_name in col_name_to_idx:
            idx = col_name_to_idx[col_name]
            letter = ws.cell(row=1, column=idx).column_letter
            dv = DataValidation(type='list', formula1=formula1, showDropDown=False)
            dv.sqref = f'{letter}2:{letter}200'
            ws.add_data_validation(dv)

    for row_data in example_rows:
        ws.append(row_data)

    return ws


def _build_data_sheet(wb, simulations, remote_labs, vr_labs, subjects=()):
    """Create a hidden _Data sheet for DB-sourced dropdowns.
    Returns (sim_formula, rlab_formula, vr_formula, subj_formula) - None for any empty list."""
    if not (simulations or remote_labs or vr_labs or subjects):
        return None, None, None, None
    ws = wb.create_sheet('_Data')
    ws.sheet_state = 'hidden'
    sim_f = rlab_f = vr_f = subj_f = None
    if simulations:
        ws['A1'] = 'Simulations'
        for i, n in enumerate(simulations, start=2):
            ws.cell(row=i, column=1, value=n)
        sim_f = f'_Data!$A$2:$A${1 + len(simulations)}'
    if remote_labs:
        ws['B1'] = 'Remote Labs'
        for i, n in enumerate(remote_labs, start=2):
            ws.cell(row=i, column=2, value=n)
        rlab_f = f'_Data!$B$2:$B${1 + len(remote_labs)}'
    if vr_labs:
        ws['C1'] = 'VR Labs'
        for i, n in enumerate(vr_labs, start=2):
            ws.cell(row=i, column=3, value=n)
        vr_f = f'_Data!$C$2:$C${1 + len(vr_labs)}'
    if subjects:
        ws['D1'] = 'Subjects'
        for i, n in enumerate(subjects, start=2):
            ws.cell(row=i, column=4, value=n)
        subj_f = f'_Data!$D$2:$D${1 + len(subjects)}'
    return sim_f, rlab_f, vr_f, subj_f


def _write_readme(ws):
    ws['A1'] = 'SCENARIO TEMPLATE - INSTRUCTIONS'
    ws['A1'].font = Font(bold=True, size=14)
    lines = [
        '',
        'SHEETS AND THEIR PURPOSE',
        '  Scenario      - One row of scenario metadata (all fields required)',
        '  Phases        - One row per phase; row order = phase order in the scenario',
        '  Activities    - One row per activity; row order = activity order within each phase',
        '  Answers       - One row per answer choice (multiple rows per activity)',
        '  Next Activity - One row per routing rule (which activity comes after which)',
        '  Evaluation    - One row per evaluatable activity (scoring groups + branching)',
        '',
        'REFERENCING RULES',
        '  Activities reference Phases by Phase Name - must match exactly.',
        '  Answers, Next Activity, and Evaluation reference Activities by Activity Name.',
        '  Activity names must be unique within the file.',
        '  Next Activity references answers by Answer Key - use the key, not the display text.',
        '  TIP: Copy-paste names and keys from other sheets to avoid typos.',
        '',
        'ACTIVITY TYPE CHOICES  (Activity Type is required)',
        '  Explanation - Present information (text, images, video) to the student',
        '  Question    - Ask a question; student selects from answer choices',
        '  Experiment  - Embed a simulation or remote lab',
        '  Guidance    - Provide targeted feedback or guidance to the student',
        '',
        'EXPERIMENT TYPE (for Experiment activities)',
        '  Select "Simulation", "Remote Lab", or "VR/AR Lab" in the Experiment Type column.',
        '  Then fill ONLY the matching name column - the other two will be greyed out:',
        '    Simulation  → fill Simulation Name only',
        '    Remote Lab  → fill Remote Lab Name only',
        '    VR/AR Lab   → fill VR Lab Name only',
        '  The name columns have a dropdown pre-loaded with resources from the platform.',
        '',
        'SUBJECTS (optional - Scenario sheet)',
        '  Use Subject 1 / Subject 2 / Subject 3 to link the scenario to platform subjects.',
        '  Each has a dropdown with all subjects from the platform.',
        '',
        'TEXT FIELDS',
        '  Activity Text supports Markdown formatting:',
        '    **bold**   _italic_   # Heading   - bullet list   [Link](https://...)',
        '  Markdown is converted to HTML automatically on import.',
        '',
        'BOOLEAN FIELDS (Is Evaluatable, Is Correct)',
        '  Use the dropdown: Yes or No (case-insensitive).',
        '',
        'ANSWER WEIGHT',
        '  Correct answers always receive weight 3 - set automatically on import.',
        '  For wrong answers choose: 1 (completely wrong) or 2 (partially correct).',
        '',
        'STUDENT PERFORMANCE CATEGORIES',
        '  After evaluation, students are placed into one of three groups:',
        '    High     - average answer weight >= 2.5',
        '    Moderate - average answer weight >= 1.5',
        '    Low      - all remaining students (fallback)',
        '  Thresholds (2.5 / 1.5 / 1.0) are fixed and applied automatically on import.',
        '  You do not need to enter them.',
        '',
        'EVALUATION SHEET - GROUPED ACTIVITIES',
        '  Use "Grouped Activity 1" through "Grouped Activity 6" to select the activities',
        '  that form this evaluation group. Include the Primary Activity itself.',
        '  Leave extra columns blank - they are ignored.',
        '',
        'SCENARIO VISIBILITY',
        '  private - visible only to you (default; can be changed after import)',
        '  org     - visible to all members of your organisation',
        '  public  - visible to all platform users',
        '',
        'NEXT ACTIVITY SHEET',
        '  Leave "Answer Key" blank for an unconditional (default) next activity.',
        '  Leave "Next Activity Name" blank to end the scenario at that activity.',
        '  Example:',
        '    Source Activity | Answer Key  | Next Activity',
        '    Welcome         |             | Quiz 1         (default route)',
        '    Quiz 1          | ans_correct | Result High    (per-answer route)',
        '    Quiz 1          | ans_wrong   | Result Low',
        '',
        'EVALUATION SHEET',
        '  Primary Activity Name: an activity with Is Evaluatable = Yes.',
        '  Grouped Activity 1-6: activities that form the scoring group (include primary).',
        '  High / Moderate / Low Performers Activity: where students go based on their score.',
        '',
        'COLUMNS MARKED WITH *',
        '  Red headers are required. Leave others blank if not needed.',
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions['A'].width = 80


def generate_blank_template(simulations=(), remote_labs=(), vr_labs=(), subjects=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'README'
    _write_readme(ws)

    sim_f, rlab_f, vr_f, subj_f = _build_data_sheet(wb, simulations, remote_labs, vr_labs, subjects)

    scen_list_cols = {
        'Language': f'"{_SCENARIO_LANGUAGES}"',
        'Visibility': '"private,org,public"',
    }
    if subj_f:
        scen_list_cols['Subject 1'] = subj_f
        scen_list_cols['Subject 2'] = subj_f
        scen_list_cols['Subject 3'] = subj_f

    _write_sheet(wb, 'Scenario', [
        ('Name', True), ('Description', True), ('Learning Goals', True),
        ('Language', True), ('Subject Domains', False),
        ('Age Min', True), ('Age Max', True),
        ('Suggested Time (min)', True), ('Visibility', True),
        ('Subject 1', False), ('Subject 2', False), ('Subject 3', False),
    ], example_rows=[
        ['My Scenario', 'A brief description', 'Students will learn...', 'English',
         '', '14', '18', '90', 'private', '', '', ''],
    ], list_cols=scen_list_cols)

    _write_sheet(wb, 'Phases', [
        ('Phase Name', True), ('Description', False),
    ], example_rows=[
        ['Engagement', ''],
        ['Hypothesis', ''],
        ['Experiment', ''],
        ['Analysis', ''],
        ['Reflection', ''],
    ])

    act_list_cols = {
        'Activity Type': '"Explanation,Question,Experiment,Guidance"',
        'Phase Name': 'Phases!$A$2:$A$200',
        'Experiment Type': '"Simulation,Remote Lab,VR/AR Lab"',
    }
    if sim_f:
        act_list_cols['Simulation Name'] = sim_f
    if rlab_f:
        act_list_cols['Remote Lab Name'] = rlab_f
    if vr_f:
        act_list_cols['VR Lab Name'] = vr_f

    acts_ws = _write_sheet(wb, 'Activities', [
        ('Activity Name', True), ('Phase Name', True), ('Text', True),
        ('Activity Type', True), ('Helper', False),
        ('Is Evaluatable', False), ('Is Primary Evaluation', False),
        ('Experiment Type', False),
        ('Simulation Name', False), ('Remote Lab Name', False), ('VR Lab Name', False),
    ], bool_cols=['Is Evaluatable', 'Is Primary Evaluation'], example_rows=[
        ['Welcome', 'Engagement', 'Welcome to this scenario! **Read carefully.**',
         'Explanation', '', 'No', 'No', '', '', '', ''],
        ['Quiz 1', 'Analysis', 'What is the boiling point of water?',
         'Question', '', 'Yes', 'Yes', '', '', '', ''],
    ], list_cols=act_list_cols)

    # Conditional formatting: grey out name columns when Experiment Type doesn't match
    # Col H=Experiment Type, I=Simulation Name, J=Remote Lab Name, K=VR Lab Name
    _grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    _grey_font = Font(color='808080')
    acts_ws.conditional_formatting.add('I2:I200',
        FormulaRule(formula=['$H2<>"Simulation"'], fill=_grey_fill, font=_grey_font))
    acts_ws.conditional_formatting.add('J2:J200',
        FormulaRule(formula=['$H2<>"Remote Lab"'], fill=_grey_fill, font=_grey_font))
    acts_ws.conditional_formatting.add('K2:K200',
        FormulaRule(formula=['$H2<>"VR/AR Lab"'], fill=_grey_fill, font=_grey_font))

    ans_ws = _write_sheet(wb, 'Answers', [
        ('Activity Name', True), ('Answer Key', True), ('Answer Text', True),
        ('Is Correct', False), ('Answer Weight', False),
    ], bool_cols=['Is Correct'], example_rows=[
        ['Quiz 1', 'ans_correct', '100°C', 'Yes', '3'],
        ['Quiz 1', 'ans_wrong',   '50°C',  'No',  '1'],
    ], list_cols={
        'Activity Name': 'Activities!$A$2:$A$200',
        'Answer Weight': '"3,2,1"',
    })
    for _row in range(4, 201):
        ans_ws.cell(row=_row, column=2, value=f'=IF(A{_row}="","","ans_"&(ROW()-1))')

    _write_sheet(wb, 'Next Activity', [
        ('Source Activity Name', True), ('Answer Key', False), ('Next Activity Name', False),
    ], example_rows=[
        ['Welcome',  '',            'Quiz 1'],
        ['Quiz 1',   'ans_correct', 'Well Done'],
        ['Quiz 1',   'ans_wrong',   'Try Again'],
    ], list_cols={
        'Source Activity Name': 'Activities!$A$2:$A$200',
        'Answer Key': 'Answers!$B$2:$B$200',
        'Next Activity Name': 'Activities!$A$2:$A$200',
    })

    eval_cols = (
        [('Primary Activity Name', True), ('Grouped Activity 1', True)]
        + [(col, False) for col in _GROUPED_ACT_COLS[1:]]
        + [
            ('High Performers Activity', False),
            ('Moderate Performers Activity', False),
            ('Low Performers Activity', False),
        ]
    )
    eval_list_cols = {col: 'Activities!$A$2:$A$200' for col in (
        ['Primary Activity Name'] + _GROUPED_ACT_COLS
        + ['High Performers Activity', 'Moderate Performers Activity', 'Low Performers Activity']
    )}
    _write_sheet(wb, 'Evaluation', eval_cols, example_rows=[
        ['Quiz 1', 'Quiz 1', 'Quiz 2', '', '', '', '',
         'Result High', 'Result Standard', 'Result Low'],
    ], list_cols=eval_list_cols)

    return wb
