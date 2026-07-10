import io
import os
import re
import uuid
import zipfile

import openpyxl
import markdown2
from django.conf import settings
from django.utils.html import strip_tags
from psycopg2.extras import NumericRange

from .models import (
    Scenario, Simulation, ExperimentLL, VRARExperiment, ActivityType, Subject,
)

# ── Column definitions ────────────────────────────────────────────────────

REQUIRED_SHEETS = ['README', 'Scenario', 'Phases', 'Activities', 'Answers', 'Next Activity', 'Evaluation']

SCENARIO_REQUIRED = ['Name']
PHASES_REQUIRED = ['Phase Name']
ACTIVITIES_REQUIRED = ['Activity Name', 'Phase Name', 'Text', 'Activity Type']
ANSWERS_REQUIRED = ['Activity Name', 'Answer Key', 'Answer Text']
ROUTING_REQUIRED = ['Source Activity Name']
EVALUATION_REQUIRED = ['Primary Activity Name', 'Grouped Activity 1']

BOOL_COLS_ACTIVITIES = ('Is Evaluatable', 'Is Primary Evaluation')

_GROUPED_ACT_COLS = [f'Grouped Activity {i}' for i in range(1, 7)]


def _grouped_acts(ev):
    """Return the list of non-empty grouped activity names from the multi-column format."""
    return [ev.get(col, '').strip() for col in _GROUPED_ACT_COLS if ev.get(col, '').strip()]
BOOL_COLS_ANSWERS = ('Is Correct',)

# Fixed performance thresholds — not editable by teachers
_SCORE_HIGH     = 2.5
_SCORE_MODERATE = 1.5
_SCORE_LOW      = 1.0


def _to_bool(value, default=False):
    v = str(value).strip().lower()
    if v == 'yes':
        return True
    if v == 'no':
        return False
    return default


def _to_float(value):
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def _to_int(value):
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return None


def _md(text):
    if not text:
        return ''
    return markdown2.markdown(str(text), extras=['fenced-code-blocks', 'tables'])


# ── Main class ────────────────────────────────────────────────────────────

class ScenarioImporter:
    def __init__(self, file_obj, user):
        self.file_obj = file_obj
        self.user = user
        self.errors = []

        self._wb_file = file_obj          # replaced with xlsx BytesIO if ZIP uploaded
        self._image_url_map = {}          # 'images/name.ext' -> '/media/tinymce/uuid.ext'
        self._prefix_images = {}          # img_prefix -> [(num, server_url), ...]

        self.scenario_data = {}
        self.phases = []
        self.activities = []
        self.activity_map = {}   # name -> dict
        self.answers = []
        self.answer_map = {}     # (activity_name, answer_key) -> dict
        self.routing = []
        self.evaluation = []

        self._simulations = {}
        self._remote_labs = {}
        self._vr_labs = {}
        self._activity_types = {}
        self._subjects = {}

    def run(self):
        """Parse, validate, create. Returns (Scenario | None, errors list)."""
        self._prepare_source()
        if not self.errors:
            self._parse()
        if not self.errors:
            self._validate()
        if not self.errors:
            try:
                return self._create(), []
            except ValueError as exc:
                return None, [{'sheet': 'Scenario', 'row': 2, 'column': 'Name', 'message': str(exc)}]
        return None, self.errors

    def _prepare_source(self):
        """Detect ZIP upload: extract xlsx + save images, build URL rewrite map."""
        # XLSX files are also ZIP archives (PK magic bytes), so distinguish by extension
        name = getattr(self.file_obj, 'name', '') or ''
        if not name.lower().endswith('.zip'):
            return  # plain xlsx — nothing to do

        try:
            raw = self.file_obj.read()
            self.file_obj.seek(0)
        except Exception:
            self._add_error('File', 0, '-', 'Cannot read ZIP file.')
            return

        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                names = zf.namelist()

                # Find xlsx — prefer root level, fall back to any level in the ZIP
                xlsx_names = [n for n in names if n.lower().endswith('.xlsx') and '/' not in n]
                if not xlsx_names:
                    xlsx_names = [n for n in names if n.lower().endswith('.xlsx')]
                if not xlsx_names:
                    self._add_error('File', 0, '-', 'No .xlsx file found in the ZIP.')
                    return
                xlsx_entry = xlsx_names[0]
                self._wb_file = io.BytesIO(zf.read(xlsx_entry))

                # Derive the images/ prefix relative to the xlsx location
                xlsx_dir = xlsx_entry.rsplit('/', 1)[0] + '/' if '/' in xlsx_entry else ''
                images_prefix = xlsx_dir + 'images/'

                # Save images and build URL map
                media_url = getattr(settings, 'MEDIA_URL', '/media/').rstrip('/')
                dest_dir = os.path.join(settings.MEDIA_ROOT, 'tinymce')
                os.makedirs(dest_dir, exist_ok=True)
                for zip_path in names:
                    if not zip_path.startswith(images_prefix) or zip_path.endswith('/'):
                        continue
                    img_name = zip_path[len(images_prefix):]
                    if '/' in img_name:
                        continue
                    ext = os.path.splitext(img_name)[1].lower() or '.jpg'
                    new_name = f'{uuid.uuid4().hex}{ext}'
                    with open(os.path.join(dest_dir, new_name), 'wb') as f:
                        f.write(zf.read(zip_path))
                    self._image_url_map[zip_path] = f'{media_url}/tinymce/{new_name}'

                # Group images by naming prefix so they can be auto-appended
                # to activities that don't yet reference them in their HTML.
                for zip_path, server_url in self._image_url_map.items():
                    img_name = zip_path.split('/')[-1]
                    m = re.match(r'^(.+)_image_(\d+)(\.\w+)?$', img_name)
                    if m:
                        prefix, num = m.group(1), int(m.group(2))
                        self._prefix_images.setdefault(prefix, []).append((num, server_url))
        except zipfile.BadZipFile:
            self._add_error('File', 0, '-', 'Cannot read ZIP file — the file may be corrupted.')
        except Exception as exc:
            self._add_error('File', 0, '-', f'Error reading ZIP: {exc}')

    def _add_error(self, sheet, row, column, message):
        self.errors.append({'sheet': sheet, 'row': row, 'column': column, 'message': message})

    # ── Parse ─────────────────────────────────────────────────────────────

    def _parse(self):
        try:
            wb = openpyxl.load_workbook(self._wb_file, read_only=True, data_only=True)
        except Exception:
            self._add_error('File', 0, '-', 'Cannot read file. Ensure it is a valid .xlsx file.')
            return

        sheet_index = {ws.title.lower(): ws.title for ws in wb.worksheets}
        missing = [s for s in REQUIRED_SHEETS if s.lower() not in sheet_index]
        for s in missing:
            self._add_error('File', 0, '-', f'Missing required sheet: "{s}"')
        if self.errors:
            return

        def ws(name):
            return wb[sheet_index[name.lower()]]

        self.scenario_data = self._parse_single_row(ws('Scenario'), 'Scenario', SCENARIO_REQUIRED)
        self.phases = self._parse_rows(ws('Phases'), 'Phases', PHASES_REQUIRED)
        self.activities = self._parse_rows(ws('Activities'), 'Activities', ACTIVITIES_REQUIRED)
        self.activity_map = {a['Activity Name']: a for a in self.activities if a.get('Activity Name')}
        self.answers = self._parse_rows(ws('Answers'), 'Answers', ANSWERS_REQUIRED)
        self.answer_map = {
            (a['Activity Name'], a['Answer Key']): a
            for a in self.answers
            if a.get('Activity Name') and a.get('Answer Key')
        }
        self.routing = self._parse_rows(ws('Next Activity'), 'Next Activity', ROUTING_REQUIRED)
        self.evaluation = self._parse_rows(ws('Evaluation'), 'Evaluation', EVALUATION_REQUIRED)

    def _read_headers(self, ws, sheet_name, required_cols):
        rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not rows or all(v is None for v in rows[0]):
            self._add_error(sheet_name, 1, '-', 'Header row is empty.')
            return None
        headers = {}
        for i, h in enumerate(rows[0]):
            if h is not None:
                headers[str(h).replace('*', '').strip()] = i
        missing = [c for c in required_cols if c not in headers]
        for c in missing:
            self._add_error(sheet_name, 1, c, f'Required column "{c}" not found in header row.')
        return None if missing else headers

    def _parse_rows(self, ws, sheet_name, required_cols):
        headers = self._read_headers(ws, sheet_name, required_cols)
        if headers is None:
            return []
        rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vals = [str(v).strip() if v is not None else '' for v in row]
            if all(v == '' for v in vals):
                continue
            d = {col: (vals[idx] if idx < len(vals) else '') for col, idx in headers.items()}
            d['_row'] = row_num
            rows.append(d)
        return rows

    def _parse_single_row(self, ws, sheet_name, required_cols):
        headers = self._read_headers(ws, sheet_name, required_cols)
        if headers is None:
            return {}
        data_rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        if not data_rows:
            return {}
        vals = [str(v).strip() if v is not None else '' for v in data_rows[0]]
        d = {col: (vals[idx] if idx < len(vals) else '') for col, idx in headers.items()}
        d['_row'] = 2
        return d

    # ── Validate ──────────────────────────────────────────────────────────

    def _validate(self):
        self._load_db_lookups()
        self._validate_scenario()
        self._validate_phases()
        self._validate_activities()
        self._validate_answers()
        self._validate_routing()
        self._validate_evaluation()

    def _load_db_lookups(self):
        self._simulations = {s.name: s for s in Simulation.objects.all()}
        self._remote_labs = {l.name: l for l in ExperimentLL.objects.all()}
        self._vr_labs = {v.name: v for v in VRARExperiment.objects.all()}
        self._activity_types = {at.name: at for at in ActivityType.objects.all()}
        self._subjects = {s.name: s for s in Subject.objects.all()}

    def _validate_scenario(self):
        if not self.scenario_data:
            self._add_error('Scenario', 2, 'Name', 'Scenario data row is missing.')
            return
        name = self.scenario_data.get('Name', '').strip()
        if not name:
            self._add_error('Scenario', 2, 'Name', 'Scenario Name is required.')
        elif Scenario.objects.filter(name=name).exists():
            self._add_error('Scenario', 2, 'Name', f'A scenario named "{name}" already exists.')
        vis = self.scenario_data.get('Visibility', '').strip().lower()
        if vis and vis not in ('private', 'org', 'public'):
            self._add_error('Scenario', 2, 'Visibility', 'Visibility must be "private", "org", or "public".')
        for col in ('Age Min', 'Age Max', 'Suggested Time (min)'):
            v = self.scenario_data.get(col, '').strip()
            if v and _to_int(v) is None:
                self._add_error('Scenario', 2, col, f'"{col}" must be a whole number.')
        for i in range(1, 4):
            col = f'Subject {i}'
            v = self.scenario_data.get(col, '').strip()
            if v and v not in self._subjects:
                self._add_error('Scenario', 2, col, f'Subject "{v}" not found in the database.')

    def _validate_phases(self):
        if not self.phases:
            self._add_error('Phases', 2, 'Phase Name', 'At least one phase is required.')

    def _validate_activities(self):
        if not self.activities:
            self._add_error('Activities', 2, 'Activity Name', 'At least one activity is required.')
            return
        phase_names = {p['Phase Name'] for p in self.phases if p.get('Phase Name')}
        seen = set()
        for a in self.activities:
            row = a['_row']
            name = a.get('Activity Name', '').strip()
            if not name:
                self._add_error('Activities', row, 'Activity Name', 'Activity Name is required.')
            elif name in seen:
                self._add_error('Activities', row, 'Activity Name', f'Duplicate activity name: "{name}".')
            else:
                seen.add(name)
            if not a.get('Text', '').strip():
                self._add_error('Activities', row, 'Text', 'Text is required.')
            phase = a.get('Phase Name', '').strip()
            if not phase:
                self._add_error('Activities', row, 'Phase Name', 'Phase Name is required.')
            elif phase not in phase_names:
                self._add_error('Activities', row, 'Phase Name', f'Phase "{phase}" not found in Phases sheet.')
            activity_type = a.get('Activity Type', '').strip()
            if not activity_type:
                self._add_error('Activities', row, 'Activity Type', 'Activity Type is required.')
            elif activity_type not in self._activity_types:
                self._add_error('Activities', row, 'Activity Type',
                                f'Activity type "{activity_type}" not found in the database.')
            for col in BOOL_COLS_ACTIVITIES:
                v = a.get(col, '').strip().lower()
                if v and v not in ('yes', 'no'):
                    self._add_error('Activities', row, col, f'"{col}" must be "Yes" or "No".')
            if (_to_bool(a.get('Is Primary Evaluation', 'No'))
                    and not _to_bool(a.get('Is Evaluatable', 'No'))):
                self._add_error('Activities', row, 'Is Primary Evaluation',
                                'Is Primary Evaluation = Yes requires Is Evaluatable = Yes.')
            for col, lookup in (
                ('Simulation Name', self._simulations),
                ('Remote Lab Name', self._remote_labs),
                ('VR Lab Name', self._vr_labs),
            ):
                v = a.get(col, '').strip()
                if v and v not in lookup:
                    self._add_error('Activities', row, col, f'"{v}" not found in the database.')

    def _validate_answers(self):
        seen_answers = set()
        for ans in self.answers:
            row = ans['_row']
            act_name = ans.get('Activity Name', '').strip()
            ans_key = ans.get('Answer Key', '').strip()
            ans_text = ans.get('Answer Text', '').strip()
            if act_name not in self.activity_map:
                self._add_error('Answers', row, 'Activity Name',
                                f'Activity "{act_name}" not found in Activities sheet.')
            if not ans_key:
                self._add_error('Answers', row, 'Answer Key', 'Answer Key is required.')
            if not ans_text:
                self._add_error('Answers', row, 'Answer Text', 'Answer Text is required.')
            pair = (act_name, ans_key)
            if pair in seen_answers:
                self._add_error('Answers', row, 'Answer Key',
                                f'Duplicate answer key "{ans_key}" for activity "{act_name}".')
            elif act_name and ans_key:
                seen_answers.add(pair)
            for col in BOOL_COLS_ANSWERS:
                v = ans.get(col, '').strip().lower()
                if v and v not in ('yes', 'no'):
                    self._add_error('Answers', row, col, f'"{col}" must be "Yes" or "No".')
            v = ans.get('Answer Weight', '').strip()
            if v and _to_int(v) is None:
                self._add_error('Answers', row, 'Answer Weight', 'Answer Weight must be a whole number.')

    def _validate_routing(self):
        seen_pairs = set()
        for r in self.routing:
            row = r['_row']
            src = r.get('Source Activity Name', '').strip()
            if not src:
                self._add_error('Next Activity', row, 'Source Activity Name',
                                'Source Activity Name is required.')
                continue
            if src not in self.activity_map:
                self._add_error('Next Activity', row, 'Source Activity Name',
                                f'Activity "{src}" not found in Activities sheet.')
                continue
            ans_key = r.get('Answer Key', '').strip()
            pair = (src, ans_key)
            if pair in seen_pairs:
                self._add_error('Next Activity', row, 'Answer Key',
                                f'Duplicate routing rule for activity "{src}" / '
                                f'answer key "{ans_key or "(default)"}".')
            seen_pairs.add(pair)
            if ans_key and (src, ans_key) not in self.answer_map:
                self._add_error('Next Activity', row, 'Answer Key',
                                f'Answer key "{ans_key}" not found for activity "{src}".')
            next_act = r.get('Next Activity Name', '').strip()
            if next_act and next_act not in self.activity_map:
                self._add_error('Next Activity', row, 'Next Activity Name',
                                f'Activity "{next_act}" not found in Activities sheet.')

    def _validate_evaluation(self):
        evaluatable = {
            a['Activity Name']
            for a in self.activities
            if _to_bool(a.get('Is Evaluatable', 'No'))
        }
        ev_seen = set()
        for ev in self.evaluation:
            row = ev['_row']
            primary = ev.get('Primary Activity Name', '').strip()
            if not primary:
                self._add_error('Evaluation', row, 'Primary Activity Name',
                                'Primary Activity Name is required.')
                continue
            if primary not in self.activity_map:
                self._add_error('Evaluation', row, 'Primary Activity Name',
                                f'Activity "{primary}" not found in Activities sheet.')
                continue
            if primary not in evaluatable:
                self._add_error('Evaluation', row, 'Primary Activity Name',
                                f'Activity "{primary}" is not marked Is Evaluatable = Yes.')
            ev_seen.add(primary)
            grouped = _grouped_acts(ev)
            if not grouped:
                self._add_error('Evaluation', row, 'Grouped Activity 1',
                                'At least one grouped activity is required.')
            else:
                for g in grouped:
                    if g not in self.activity_map:
                        self._add_error('Evaluation', row, 'Grouped Activity 1',
                                        f'Activity "{g}" not found in Activities sheet.')
            for col in ('High Performers Activity', 'Moderate Performers Activity', 'Low Performers Activity'):
                v = ev.get(col, '').strip()
                if v and v not in self.activity_map:
                    self._add_error('Evaluation', row, col,
                                    f'Activity "{v}" not found in Activities sheet.')
        for name in evaluatable:
            if name not in ev_seen:
                a = self.activity_map[name]
                self._add_error('Activities', a.get('_row', '?'), 'Is Evaluatable',
                                f'Activity "{name}" is marked Is Evaluatable = Yes '
                                f'but has no row in the Evaluation sheet.')

    def _create(self):
        from django.db import transaction, IntegrityError
        from django.utils.html import strip_tags
        from .models import (
            Phase, Activity, Answer,
            NextQuestionLogic, QuestionBunch, EvQuestionBranching,
        )

        with transaction.atomic():
            # 1. Scenario
            age_min = _to_int(self.scenario_data.get('Age Min', ''))
            age_max = _to_int(self.scenario_data.get('Age Max', ''))
            age_range = NumericRange(age_min, age_max) if (age_min is not None and age_max is not None) else None
            try:
                scenario = Scenario.objects.create(
                    name=self.scenario_data['Name'].strip(),
                    description=self.scenario_data.get('Description', '').strip(),
                    learning_goals=self.scenario_data.get('Learning Goals', '').strip(),
                    language=self.scenario_data.get('Language', ''),
                    subject_domains=self.scenario_data.get('Subject Domains', ''),
                    age_of_students=age_range,
                    suggested_learning_time=_to_int(self.scenario_data.get('Suggested Time (min)', '')) or None,
                    visibility_status=self.scenario_data.get('Visibility', '').strip().lower() or 'private',
                    created_by=self.user,
                    updated_by=self.user,
                )
            except IntegrityError:
                raise ValueError(
                    f'A scenario named "{self.scenario_data["Name"].strip()}" already exists (created concurrently).'
                )

            # 1b. Subjects M2M
            subject_objs = [
                self._subjects[self.scenario_data.get(f'Subject {i}', '').strip()]
                for i in range(1, 4)
                if self.scenario_data.get(f'Subject {i}', '').strip() in self._subjects
            ]
            if subject_objs:
                scenario.subjects.set(subject_objs)

            # 2. Phases (row order preserved)
            phase_obj_map = {}
            for ph in self.phases:
                obj = Phase.objects.create(
                    name=ph['Phase Name'],
                    description=ph.get('Description', ''),
                    scenario=scenario,
                    created_by=self.user,
                    updated_by=self.user,
                )
                phase_obj_map[ph['Phase Name']] = obj

            # 3. Activities (row order preserved)
            activity_obj_map = {}
            for act in self.activities:
                html_col = act.get('Text HTML', '').strip()
                if html_col:
                    # Restore image URLs from ZIP-relative paths to server media URLs
                    html = html_col
                    for zip_path, server_url in self._image_url_map.items():
                        html = html.replace(f'src="{zip_path}"', f'src="{server_url}"')
                else:
                    html = _md(act.get('Text', ''))

                # Auto-append images placed in the ZIP under this activity's
                # naming prefix that are not already referenced in the HTML.
                phase_slug = re.sub(r'[^\w]+', '_', (act.get('Phase Name', '') or '').strip()).strip('_').lower()
                act_slug = re.sub(r'[^\w]+', '_', (act['Activity Name'] or '').strip()).strip('_').lower()
                img_prefix = f'phase_{phase_slug}_activity_{act_slug}'
                for _num, server_url in sorted(self._prefix_images.get(img_prefix, [])):
                    if server_url not in html:
                        html = html.rstrip() + f'<p><img src="{server_url}"></p>'
                obj = Activity.objects.create(
                    name=act['Activity Name'],
                    text=html,
                    plain_text=strip_tags(html),
                    is_evaluatable=_to_bool(act.get('Is Evaluatable', 'No')),
                    is_primary_ev=_to_bool(act.get('Is Primary Evaluation', 'No')),
                    helper=act.get('Helper', ''),
                    scenario=scenario,
                    phase=phase_obj_map[act['Phase Name']],
                    activity_type=self._activity_types.get(act.get('Activity Type', '').strip()),
                    simulation=self._simulations.get(act.get('Simulation Name', '').strip()),
                    experiment_ll=self._remote_labs.get(act.get('Remote Lab Name', '').strip()),
                    vr_ar_experiment=self._vr_labs.get(act.get('VR Lab Name', '').strip()),
                    created_by=self.user,
                    updated_by=self.user,
                )
                activity_obj_map[act['Activity Name']] = obj
                # compound key so same name in different phases resolves correctly
                phase_name = act.get('Phase Name', '').strip()
                if phase_name:
                    activity_obj_map[(phase_name, act['Activity Name'])] = obj

            def _act(name, phase=''):
                """Look up an activity by (phase, name), falling back to name only."""
                name = (name or '').strip()
                phase = (phase or '').strip()
                if phase and (phase, name) in activity_obj_map:
                    return activity_obj_map[(phase, name)]
                return activity_obj_map.get(name)

            # 4. Answers  (keyed by Answer Key for routing lookups)
            answer_obj_map = {}
            for ans in self.answers:
                is_correct_bool = _to_bool(ans.get('Is Correct', 'No'))
                if is_correct_bool:
                    weight = 3
                else:
                    weight = _to_int(ans.get('Answer Weight', '')) or 1
                ans_obj = Answer.objects.create(
                    activity=activity_obj_map[ans['Activity Name']],
                    text=ans.get('Answer Text', '').strip(),
                    is_correct=is_correct_bool,
                    answer_weight=weight,
                    created_by=self.user,
                    updated_by=self.user,
                )
                answer_obj_map[(ans['Activity Name'], ans['Answer Key'])] = ans_obj

            # 5. NextQuestionLogic
            for r in self.routing:
                src_name = r['Source Activity Name'].strip()
                src_phase = r.get('Source Phase', '').strip()
                src = _act(src_name, src_phase)
                ans_key = r.get('Answer Key', '').strip()
                ans_obj = (
                    answer_obj_map.get((src_name, ans_key))
                    if ans_key else None
                )
                next_name = r.get('Next Activity Name', '').strip()
                next_phase = r.get('Next Phase', '').strip()
                NextQuestionLogic.objects.create(
                    activity=src,
                    answer=ans_obj,
                    next_activity=_act(next_name, next_phase) if next_name else None,
                )

            # 6. QuestionBunch + EvQuestionBranching with fixed thresholds
            for ev in self.evaluation:
                primary = _act(ev['Primary Activity Name'])
                grouped = _grouped_acts(ev)
                grouped_ids = [_act(n).id for n in grouped if _act(n)]
                QuestionBunch.objects.create(
                    activity_primary=primary,
                    activity_ids=grouped_ids,
                )

                high_name = ev.get('High Performers Activity', '').strip()
                mod_name  = ev.get('Moderate Performers Activity', '').strip()
                low_name  = ev.get('Low Performers Activity', '').strip()

                high_act = _act(high_name)
                mod_act  = _act(mod_name)
                low_act  = _act(low_name)

                EvQuestionBranching.objects.create(
                    activity=primary,
                    next_question_on_high=high_act,
                    next_question_on_mid=mod_act,
                    next_question_on_low=low_act,
                )

                # Apply fixed thresholds to branch target activities
                if high_act:
                    Activity.objects.filter(pk=high_act.pk).update(score_limit=_SCORE_HIGH)
                if mod_act:
                    Activity.objects.filter(pk=mod_act.pk).update(score_limit=_SCORE_MODERATE)
                if low_act:
                    Activity.objects.filter(pk=low_act.pk).update(score_limit=_SCORE_LOW)

        return scenario
