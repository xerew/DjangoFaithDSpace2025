"""Downloadable organization reports."""

from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .statistics import get_organization_statistics


BLUE = '1A56DB'
DARK_BLUE = '012970'
LIGHT_BLUE = 'EAF0FF'
WHITE = 'FFFFFF'


def _style_title(cell):
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.fill = PatternFill('solid', fgColor=DARK_BLUE)
    cell.alignment = Alignment(vertical='center')


def _style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=BLUE)
        cell.alignment = Alignment(vertical='center')


def _fit_columns(worksheet, minimum=12, maximum=46):
    for column_cells in worksheet.columns:
        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        ) + 2
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(
            minimum, min(width, maximum)
        )


def _add_table_sheet(workbook, title, headers, rows):
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    _style_header(worksheet[1])
    for row in rows:
        worksheet.append(row)
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions
    _fit_columns(worksheet)
    return worksheet


def build_organization_report(organization):
    """Return an in-memory XLSX report for ``organization``."""
    stats = get_organization_statistics(organization, include_group_rows=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Summary'
    summary.merge_cells('A1:D1')
    summary['A1'] = f'{organization.name} — Organization Data Report'
    _style_title(summary['A1'])
    summary.row_dimensions[1].height = 30

    generated_at = timezone.localtime()
    summary.append(['Generated', generated_at.strftime('%Y-%m-%d %H:%M %Z')])
    summary.append(['Organization', organization.name])
    summary.append(['Short name', organization.short_name])
    summary.append([])
    summary.append(['Metric', 'Value'])
    _style_header(summary[6])
    metrics = [
        ('Teachers', stats['teacher_count']),
        ('Student groups', stats['group_count']),
        ('Registered students', stats['student_count']),
        ('Active students', stats['active_student_count']),
        ('Implementations', stats['implementation_count']),
        ('Scenarios used', stats['scenario_count']),
        ('Scenarios assigned', stats['assigned_scenario_count']),
    ]
    for metric in metrics:
        summary.append(metric)

    summary.append([])
    summary.append(['How the figures are calculated'])
    summary['A15'].font = Font(bold=True, color=DARK_BLUE)
    summary.merge_cells('A16:D18')
    summary['A16'] = (
        'Teachers are current organization members with the Teacher role. '
        'Students are members of student groups created by those teachers. '
        'One implementation is one distinct student–scenario record. '
        'Individual student names are not included in this report.'
    )
    summary['A16'].alignment = Alignment(wrap_text=True, vertical='top')
    _fit_columns(summary)
    summary.column_dimensions['A'].width = 30
    summary.column_dimensions['B'].width = 24

    _add_table_sheet(
        workbook,
        'Scenario Usage',
        ['Scenario', 'Implementations', 'Visibility', 'Language'],
        [
            [
                row['scenario__name'],
                row['implementations'],
                (row['scenario__visibility_status'] or 'unspecified').title(),
                row['scenario__language'] or 'Unspecified',
            ]
            for row in stats['scenario_usage']
        ],
    )
    _add_table_sheet(
        workbook,
        'Scenario Categories',
        ['Category', 'Scenarios'],
        [[row['category'], row['scenario_count']] for row in stats['category_rows']],
    )
    _add_table_sheet(
        workbook,
        'Activity Types',
        ['Activity type', 'Scenarios', 'Activities'],
        [
            [row['name'], row['scenario_count'], row['activity_count']]
            for row in stats['activity_type_rows']
        ],
    )
    _add_table_sheet(
        workbook,
        'Student Groups',
        ['Group', 'Teacher', 'Students', 'Assigned scenarios', 'Implementations'],
        [
            [
                row['name'],
                row['teacher'],
                row['students'],
                row['assigned_scenarios'],
                row['implementations'],
            ]
            for row in stats['group_rows']
        ],
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
